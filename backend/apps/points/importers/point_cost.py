from openpyxl import load_workbook

from apps.points.models import Point, PointCost


class PointCostImporter(object):
    def __init__(self, file):
        self.workbook = load_workbook(filename=file)
        self.data = []
        self.costs = []

    def import_costs(self):
        self.collect_data()
        self.prepare_points_data()
        self.create_points_costs()

    def collect_data(self):
        for sheet in self.workbook.worksheets:
            for row in range(PointCost.START_ROW, sheet.max_row + 1):
                row_data = []
                if not sheet.cell(row=row, column=20).value:
                    continue
                for col in range(1, sheet.max_column + 1):
                    row_data.append(sheet.cell(row=row, column=col).value)
                self.data.append(row_data)

    def create_points_costs(self):
        for item in self.costs:
            point = item.pop('point')
            PointCost.objects.get_or_create(**item, point=point)

    def prepare_points_data(self):
        for row in self.data:
            p_obj = Point.objects.filter(ppe_number=row[20]).first()
            if p_obj:
                point = {
                    'point': p_obj,
                }
                for key, index in PointCost.IMPORT_MAPPER.items():
                    point[key] = row[index]
                self.costs.append(point)
