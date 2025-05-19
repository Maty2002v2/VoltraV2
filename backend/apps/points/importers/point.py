from datetime import datetime

from openpyxl import load_workbook

from apps.division.models import Division
from apps.points.models import Point


class PointImporter(object):
    def __init__(self, file):
        self.workbook = load_workbook(filename=file)
        self.data = []
        self.points = []

    def import_points(self):
        self.collect_data()
        self.prepare_points_data()
        self.create_points()

    def collect_data(self):
        for sheet in self.workbook.worksheets:
            for row in range(Point.START_ROW, sheet.max_row + 1):
                row_data = []
                if not sheet.cell(row=row, column=1).value:
                    continue
                for col in range(1, sheet.max_column + 1):
                    row_data.append(sheet.cell(row=row, column=col).value)
                self.data.append(row_data)

    def create_division(self, point):
        division_keys = {}
        for key in list(point.keys()):
            if key.startswith('division'):
                division_keys[key.replace('division_', '')] = point.pop(key)

        division_obj, _ = Division.objects.get_or_create(**division_keys)
        return division_obj

    def set_proper_date_format(self, point_data):
        date_fields = {'sale_start', 'sale_end', 'termination_date'}
        for key, value in point_data.items():
            if key in date_fields:
                if isinstance(value, datetime):
                    point_data[key] = value.strftime('%Y-%m-%d')
                elif isinstance(value, str):
                    point_data[key] = datetime.strptime(value, '%d.%m.%Y').strftime('%Y-%m-%d')

    def create_points(self):
        for point in self.points:
            division = self.create_division(point)
            point_data = {key: value or '' for key, value in point.items()}
            self.set_proper_date_format(point_data)
            Point.objects.get_or_create(**point_data, division=division, verified=True)

    def prepare_points_data(self):
        for row in self.data:
            point = {}
            for key, index in Point.IMPORT_MAPPER.items():
                point[key] = row[index]
            self.points.append(point)
