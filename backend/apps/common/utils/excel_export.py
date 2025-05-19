from datetime import datetime

from django.contrib import messages
from django.http import HttpResponse
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
from openpyxl.writer.excel import save_virtual_workbook

STYLE = 'thin'

border = Border(
    left=Side(style=STYLE),
    right=Side(style=STYLE),
    top=Side(style=STYLE),
    bottom=Side(style=STYLE),
)


def export(request, file_path, file_name, queryset, start_row, index_column=True):  # noqa: WPS211
    workbook = load_workbook(filename=file_path)
    ws = workbook.worksheets[0]

    try:
        for row_index, obj in enumerate(queryset, start_row + 1):
            if index_column:
                ws.cell(row=row_index, column=1).value = row_index - start_row
                ws.cell(row=row_index, column=1).border = border
                start_column = 2
            else:
                start_column = 1

            for col_index, field in enumerate(obj.to_excel(), start_column):
                ws.cell(row=row_index, column=col_index).value = field
                ws.cell(row=row_index, column=col_index).border = border
        messages.add_message(request, messages.INFO, 'Eksport zakończony sukcesem.')
    except Exception as error:
        messages.add_message(request, messages.ERROR, repr(error))

    response = HttpResponse(content=save_virtual_workbook(workbook), content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="{file_name}_{date}.xls"'.format(
        file_name=file_name,
        date=datetime.today().date(),
    )
    return response
