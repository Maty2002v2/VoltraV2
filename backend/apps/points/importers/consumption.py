from datetime import datetime
from apps.points.models import Point
from openpyxl import load_workbook


class ConsumptionImporter(object):

    def __init__(self, file):
        self.workbook = load_workbook(filename=file)
        self.data = []
        self.costs = []

    def import_consumption(self):
        self.collect_data()
        self.update_points()

    def collect_data(self):
        for sheet in self.workbook.worksheets:
            for row in range(1, sheet.max_row + 1):
                lp = sheet.cell(row=row, column=1).value
                ppe_number = sheet.cell(row=row, column=2).value
                annual_consumption = sheet.cell(row=row, column=3).value
                if not lp:
                    continue
                row_data = {
                    'ppe_number': ppe_number,
                    'annual_consumption': annual_consumption
                }
                self.data.append(row_data)


    def update_points(self):
        for point in self.data:
            ppe_number = point.get('ppe_number')
            annual_consumption = point.get('annual_consumption')

            if ppe_number and annual_consumption is not None:
                Point.objects.filter(ppe_number=ppe_number).update(annual_consumption=annual_consumption, verified=True)



